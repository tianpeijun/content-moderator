#!/usr/bin/env python3
"""Generate the batch test Excel template file."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "测试用例"

headers = ["序号", "内容文本", "图片URL", "期望结果", "业务类型", "备注"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

rows = [
    [1, "这个产品质量很好，推荐购买！", "", "pass", "商品评论", "中文正常好评"],
    [2, "Great product, highly recommended!", "", "pass", "商品评论", "英文正常评论"],
    [3, "この商品はとても良いです。おすすめします。", "", "pass", "商品评论", "日文正常评论"],
    [4, "이 제품은 정말 좋습니다. 추천합니다.", "", "pass", "商品评论", "韩文正常评论"],
    [5, "这个卖家是骗子，加我微信xxx维权", "", "reject", "商品评论", "含个人联系方式"],
    [6, "Buy now! Contact WhatsApp +1234567890", "", "reject", "商品评论", "英文广告+隐私泄露"],
    [7, "商品外观很漂亮", "https://example.com/product.jpg", "pass", "商品评论", "图文混合正常评论"],
    [8, "看图片里的联系方式加我", "https://example.com/ad.jpg", "reject", "商品评论", "图片含广告信息"],
    [9, "产品一般般，不太推荐", "", "pass", "商品评论", "负面但合规评论"],
    [10, "이 판매자는 사기꾼입니다!", "", "flag", "商品评论", "韩文欺诈指控"],
]

for row in rows:
    ws.append(row)

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 45
ws.column_dimensions["C"].width = 40
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 15
ws.column_dimensions["F"].width = 25

wb.save("docs/batch-test-template.xlsx")
print("Template saved to docs/batch-test-template.xlsx")
