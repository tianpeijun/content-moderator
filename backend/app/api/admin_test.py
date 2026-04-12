"""Admin Batch Test API router.

Provides endpoints for batch test suite management:
- POST /api/admin/test-suites/upload       — upload & parse .xlsx test suite
- POST /api/admin/test-suites/{id}/run     — start batch test via SQS
- GET  /api/admin/test-suites/{id}/progress — query test progress
- GET  /api/admin/test-suites/{id}/report   — get test report
- POST /api/admin/test-suites/{id}/export   — export test report
- GET  /api/admin/test-records              — list historical test records
- POST /api/admin/test-records/compare      — compare two test records

Validates: Requirements 6.1, 6.2, 6.3, 6.4, 6.5, 6.6, 6.7
"""

from __future__ import annotations

import io
import json
import uuid
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from backend.app.core.auth import verify_cognito_token
from backend.app.core.config import settings
from backend.app.core.database import get_db
from backend.app.models.test_records import TestRecord
from backend.app.models.test_suites import TestSuite
from backend.app.schemas.test_records import TestRecordCompareRequest, TestRecordResponse
from backend.app.schemas.test_suites import (
    TestReportResponse,
    TestSuiteProgressResponse,
    TestSuiteUploadResponse,
)

router = APIRouter(
    prefix="/api/admin",
    dependencies=[Depends(verify_cognito_token)],
)

# Required columns in the uploaded .xlsx file
REQUIRED_COLUMNS = {"序号", "内容文本", "图片URL", "期望结果", "业务类型", "备注"}
VALID_EXPECTED_RESULTS = {"pass", "reject", "review", "flag"}


# ---------------------------------------------------------------------------
# Schemas for request bodies
# ---------------------------------------------------------------------------

from pydantic import BaseModel, Field


class RunTestRequest(BaseModel):
    """Request body for starting a batch test run."""
    rule_ids: list[str] = Field(description="要使用的规则 ID 列表")


class CompareResponse(BaseModel):
    """Response for comparing two test records."""
    record_a: TestRecordResponse
    record_b: TestRecordResponse


# ---------------------------------------------------------------------------
# POST /api/admin/test-suites/upload  (Requirement 6.1, 6.2)
# ---------------------------------------------------------------------------

@router.post(
    "/test-suites/upload",
    response_model=TestSuiteUploadResponse,
    status_code=status.HTTP_201_CREATED,
    summary="上传测试集",
)
async def upload_test_suite(
    file: UploadFile = File(..., description="Excel (.xlsx) 测试集文件"),
    db: Session = Depends(get_db),
    _claims: dict = Depends(verify_cognito_token),
) -> TestSuiteUploadResponse:
    """Upload and parse an .xlsx test suite file."""
    # Validate file extension
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="仅支持 .xlsx 格式的文件",
        )

    contents = await file.read()

    # Parse workbook
    try:
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="无法解析 Excel 文件，请确认文件格式正确",
        )

    ws = wb.active
    if ws is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel 文件中没有活动工作表",
        )

    # Read header row
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel 文件为空",
        )

    headers = {str(h).strip() if h else "" for h in rows[0]}
    missing = REQUIRED_COLUMNS - headers
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"缺少必要列: {', '.join(sorted(missing))}",
        )

    # Map header names to column indices
    header_list = [str(h).strip() if h else "" for h in rows[0]]
    expected_result_idx = header_list.index("期望结果")

    # Validate data rows
    data_rows = rows[1:]
    invalid_rows: list[str] = []
    for i, row in enumerate(data_rows, start=2):
        val = str(row[expected_result_idx]).strip().lower() if row[expected_result_idx] else ""
        if val not in VALID_EXPECTED_RESULTS:
            invalid_rows.append(f"第{i}行期望结果值无效: '{val}'")

    if invalid_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"数据验证失败: {'; '.join(invalid_rows[:10])}",
        )

    total_cases = len(data_rows)

    # Mock S3 upload — store file_key as a placeholder
    file_key = f"test-suites/{uuid.uuid4()}/{file.filename}"

    # Create DB record
    suite = TestSuite(
        name=file.filename,
        file_key=file_key,
        total_cases=total_cases,
    )
    db.add(suite)
    db.commit()
    db.refresh(suite)

    return TestSuiteUploadResponse.model_validate(suite)


# ---------------------------------------------------------------------------
# POST /api/admin/test-suites/{id}/run  (Requirement 6.3)
# ---------------------------------------------------------------------------

@router.post(
    "/test-suites/{suite_id}/run",
    status_code=status.HTTP_201_CREATED,
    summary="启动批量测试",
)
async def run_test_suite(
    suite_id: uuid.UUID,
    body: RunTestRequest,
    db: Session = Depends(get_db),
    _claims: dict = Depends(verify_cognito_token),
) -> dict[str, Any]:
    """Create a test record and send an SQS message to start batch testing."""
    suite = db.execute(
        select(TestSuite).where(TestSuite.id == suite_id)
    ).scalar_one_or_none()

    if suite is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"测试集 '{suite_id}' 不存在",
        )

    record = TestRecord(
        test_suite_id=suite.id,
        rule_ids=body.rule_ids,
        status="pending",
        progress_current=0,
        progress_total=suite.total_cases,
    )
    db.add(record)
    db.commit()
    db.refresh(record)

    # Mock SQS send — in production this would use boto3
    # sqs_client.send_message(QueueUrl=settings.sqs_batch_test_queue_url, ...)
    _mock_sqs_message = {
        "test_record_id": str(record.id),
        "test_suite_id": str(suite.id),
        "rule_ids": body.rule_ids,
        "queue_url": settings.sqs_batch_test_queue_url,
    }

    return {"test_record_id": str(record.id), "status": "pending"}


# ---------------------------------------------------------------------------
# GET /api/admin/test-suites/{id}/progress  (Requirement 6.3)
# ---------------------------------------------------------------------------

@router.get(
    "/test-suites/{suite_id}/progress",
    response_model=TestSuiteProgressResponse,
    summary="查询测试进度",
)
async def get_test_progress(
    suite_id: uuid.UUID,
    db: Session = Depends(get_db),
    _claims: dict = Depends(verify_cognito_token),
) -> TestSuiteProgressResponse:
    """Query the latest test record progress for a test suite."""
    record = db.execute(
        select(TestRecord)
        .where(TestRecord.test_suite_id == suite_id)
        .order_by(TestRecord.started_at.desc().nulls_last())
        .limit(1)
    ).scalar_one_or_none()

    if record is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"测试集 '{suite_id}' 没有测试记录",
        )

    return TestSuiteProgressResponse(
        test_record_id=record.id,
        status=record.status,
        progress_current=record.progress_current,
        progress_total=record.progress_total,
    )


# ---------------------------------------------------------------------------
# GET /api/admin/test-suites/{id}/report  (Requirement 6.4)
# ---------------------------------------------------------------------------

@router.get(
    "/test-suites/{suite_id}/report",
    response_model=TestReportResponse,
    summary="获取测试报告",
)
async def get_test_report(
    suite_id: uuid.UUID,
    db: Session = Depends(get_db),
    _claims: dict = Depends(verify_cognito_token),
) -> TestReportResponse:
    """Get the test report for the latest completed test record."""
    record = db.execute(
        select(TestRecord)
        .where(TestRecord.test_suite_id == suite_id)
        .order_by(TestRecord.started_at.desc().nulls_last())
        .limit(1)
    ).scalar_one_or_none()

    if record is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"测试集 '{suite_id}' 没有测试记录",
        )

    report = record.report or {}
    return TestReportResponse(
        test_record_id=record.id,
        test_suite_id=record.test_suite_id,
        status=record.status,
        accuracy=report.get("accuracy"),
        recall=report.get("recall"),
        f1_score=report.get("f1_score"),
        confusion_matrix=report.get("confusion_matrix"),
        error_cases=report.get("error_cases"),
        rule_hit_distribution=report.get("rule_hit_distribution"),
        started_at=record.started_at,
        completed_at=record.completed_at,
    )


# ---------------------------------------------------------------------------
# POST /api/admin/test-suites/{id}/export  (Requirement 6.5)
# ---------------------------------------------------------------------------

@router.post(
    "/test-suites/{suite_id}/export",
    summary="导出测试报告",
)
async def export_test_report(
    suite_id: uuid.UUID,
    db: Session = Depends(get_db),
    _claims: dict = Depends(verify_cognito_token),
) -> dict[str, Any]:
    """Export the test report as JSON (simplified — real export would produce .xlsx)."""
    record = db.execute(
        select(TestRecord)
        .where(TestRecord.test_suite_id == suite_id)
        .order_by(TestRecord.started_at.desc().nulls_last())
        .limit(1)
    ).scalar_one_or_none()

    if record is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"测试集 '{suite_id}' 没有测试记录",
        )

    return {
        "test_record_id": str(record.id),
        "test_suite_id": str(record.test_suite_id),
        "status": record.status,
        "report": record.report,
        "started_at": str(record.started_at) if record.started_at else None,
        "completed_at": str(record.completed_at) if record.completed_at else None,
    }


# ---------------------------------------------------------------------------
# GET /api/admin/test-records  (Requirement 6.6)
# ---------------------------------------------------------------------------

@router.get(
    "/test-records",
    response_model=list[TestRecordResponse],
    summary="历史测试记录",
)
async def list_test_records(
    page: int = Query(default=1, ge=1, description="页码"),
    page_size: int = Query(default=20, ge=1, le=100, description="每页数量"),
    db: Session = Depends(get_db),
    _claims: dict = Depends(verify_cognito_token),
) -> list[TestRecordResponse]:
    """List historical test records with pagination."""
    stmt = (
        select(TestRecord)
        .order_by(TestRecord.started_at.desc().nulls_last())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    records = db.execute(stmt).scalars().all()
    return [TestRecordResponse.model_validate(r) for r in records]


# ---------------------------------------------------------------------------
# POST /api/admin/test-records/compare  (Requirement 6.6)
# ---------------------------------------------------------------------------

@router.post(
    "/test-records/compare",
    response_model=CompareResponse,
    summary="测试记录对比",
)
async def compare_test_records(
    body: TestRecordCompareRequest,
    db: Session = Depends(get_db),
    _claims: dict = Depends(verify_cognito_token),
) -> CompareResponse:
    """Compare two test records side by side."""
    record_a = db.execute(
        select(TestRecord).where(TestRecord.id == body.record_id_a)
    ).scalar_one_or_none()
    record_b = db.execute(
        select(TestRecord).where(TestRecord.id == body.record_id_b)
    ).scalar_one_or_none()

    if record_a is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"测试记录 '{body.record_id_a}' 不存在",
        )
    if record_b is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"测试记录 '{body.record_id_b}' 不存在",
        )

    return CompareResponse(
        record_a=TestRecordResponse.model_validate(record_a),
        record_b=TestRecordResponse.model_validate(record_b),
    )
