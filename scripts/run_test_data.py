#!/usr/bin/env python3
"""
Run the 100 test cases from 评论审核测试数据.xlsx through the live API,
then write results back into the Excel file.

Columns filled:
  - API文案审核结果 (col Q, index 16)
  - API图片审核结果 (col R, index 17)
  - 文案是否一致   (col S, index 18)
  - 图片是否一致   (col T, index 19)
"""

import json
import os
import time
import sys
import requests
from openpyxl import load_workbook

API = os.environ.get("API_URL", "https://YOUR_API_GATEWAY_ID.execute-api.us-east-1.amazonaws.com/prod")
KEY = os.environ.get("API_KEY", "YOUR_API_KEY_HERE")
XLSX = "docs/评论审核测试数据.xlsx"
TIMEOUT = 120  # seconds per request

def call_moderate(text, image_url, business_type="商品评论"):
    """Call the moderation API and return (text_label, image_label) or error."""
    payload = {}
    if text:
        payload["text"] = str(text)
    if image_url:
        payload["image_url"] = str(image_url)
    if not payload:
        return ("error", "error")
    payload["business_type"] = business_type

    try:
        resp = requests.post(
            f"{API}/api/v1/moderate",
            json=payload,
            headers={"X-API-Key": KEY, "Content-Type": "application/json"},
            timeout=TIMEOUT,
        )
        if resp.status_code == 200:
            data = resp.json()
            return (
                data.get("text_label", ""),
                data.get("image_label", ""),
            )
        else:
            print(f"  HTTP {resp.status_code}: {resp.text[:200]}")
            return ("error", "error")
    except Exception as e:
        print(f"  Exception: {e}")
        return ("error", "error")


def main():
    wb = load_workbook(XLSX)
    ws = wb["评论测试数据"]

    # Verify headers
    headers = [str(c.value or "").strip() for c in ws[1]]
    print(f"Headers: {headers}")

    # Find column indices (0-based)
    col_map = {h: i for i, h in enumerate(headers)}
    idx_text = col_map.get("评论内容", 4)
    idx_img1 = col_map.get("图片1", 6)
    idx_expected_text = col_map.get("预期文案标签", 13)
    idx_expected_img = col_map.get("预期图片标签", 14)
    idx_scene = col_map.get("测试场景", 15)
    idx_api_text = col_map.get("API文案审核结果", 16)
    idx_api_img = col_map.get("API图片审核结果", 17)
    idx_text_match = col_map.get("文案是否一致", 18)
    idx_img_match = col_map.get("图片是否一致", 19)

    total = ws.max_row - 1  # exclude header
    passed_text = 0
    passed_img = 0
    processed = 0

    print(f"\nTotal test cases: {total}")
    print("=" * 60)

    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c+1).value for c in range(ws.max_column)]

        seq = row[0]
        text = row[idx_text]
        img1 = row[idx_img1]
        expected_text_label = str(row[idx_expected_text] or "").strip()
        expected_img_label = str(row[idx_expected_img] or "").strip()

        print(f"[{seq}/{total}] text={str(text)[:40]}... img={'yes' if img1 else 'no'} expected=({expected_text_label}, {expected_img_label})")

        # Call API
        actual_text_label, actual_img_label = call_moderate(text, img1)

        # Compare (normalize: API returns "none" but Excel expects "无")
        norm_actual_img = "无" if actual_img_label == "none" else actual_img_label
        text_match = "是" if actual_text_label == expected_text_label else "否"
        img_match = "是" if norm_actual_img == expected_img_label else "否"

        if text_match == "是":
            passed_text += 1
        if img_match == "是":
            passed_img += 1
        processed += 1

        # Write back to Excel (1-based column)
        ws.cell(row=row_idx, column=idx_api_text + 1, value=actual_text_label)
        ws.cell(row=row_idx, column=idx_api_img + 1, value=norm_actual_img)
        ws.cell(row=row_idx, column=idx_text_match + 1, value=text_match)
        ws.cell(row=row_idx, column=idx_img_match + 1, value=img_match)

        print(f"  -> API=({actual_text_label}, {actual_img_label}) text_match={text_match} img_match={img_match}")

        # Save every 10 rows in case of interruption
        if processed % 10 == 0:
            wb.save(XLSX)
            print(f"  [Saved progress: {processed}/{total}]")

        # Small delay to avoid throttling
        time.sleep(0.5)

    # Final save
    wb.save(XLSX)

    print("\n" + "=" * 60)
    print(f"Completed: {processed}/{total}")
    print(f"文案标签准确率: {passed_text}/{processed} = {passed_text/processed*100:.1f}%")
    print(f"图片标签准确率: {passed_img}/{processed} = {passed_img/processed*100:.1f}%")
    print(f"Results saved to {XLSX}")


if __name__ == "__main__":
    main()
